编程要素

	1.算法
		数理逻辑：数列 矩阵运算
		优化算法：贪心算法 递归  动态规则 回溯
		搜索
		排序
		模式匹配
	
	2.数据结构
		序列类型：列表 集合 堆栈
		
		键值类型：哈希表  Hashtable 类实现一个哈希表，该哈希表将键映射到相应的值。任何非 null 对象都可以用作键或值。为了成功地在哈希表中存储和获取对象，用作键的对象必须实现 hashCode 方法和 equals 方法。
				 map键值类型
		
		字符串
		
		树： 它是由n（n>=1）个有限节点组成一个具有层次关系的集合  树形数据结构可以表示数据表素之间一对多的关系  以树与二叉树最为常用
			堆：在计算机科学中，堆是一种特殊的树形数据结构，每个结点都有一个值。通常我们所说的堆的数据结构，是指二叉堆。堆的特点是根结点的值最小（或最大），且根结点的两个子树也是一个堆。 左偏树(Leftist Tree)是一种可并堆的实现
		
		序列化：xml  json protobuf
		
		图论~
	
	3.设计模式
		SOLID:
			SRP	 单一责任原则
			OCP	 开放封闭原则
			LSP	 里氏替换原则
			ISP  接口分离原则
			DIP	 依赖倒置原则
				
				解释：【单一职责原则】：职责单一，如果设计两个功能就拆分。你不愿意设计一个关于壶的类，即当夜壶又当茶壶吧。这个时候，你自然就会设计成夜壶类和茶壶类。【开放闭合原则】:对扩展开放，对修改闭合。想改变家族基因，就生孩子去。在自己身上动刀子，该不了基因。就算改了工作量也大，也不知道修改后能生下来个什么样的怪胎【里氏替换原则】：子类不要试图跟父类表现不一致。在古代，一个家族都是一个姓，带带相传，如果想要改姓，那你就离开我家，不能上我祠堂啦。或者，一个将类是男人，继承的都是男人长小鸡鸡，男人长喉结，你突然搞个子类，男人生孩子，这就不合适啦。【接口分离原则】：不要设计大一统的接口，要细化。比如你要设计一个吃饭的接口，就直接给设计成俩，吃米和吃面。【依赖反转原则】：抽象不要依赖具体。比如我要吃饭，不要强制要求我适用金碗，咋地，没金碗就看着饭饿死呀。我随便找个碗就行，实在没碗，我直接用手抓就行。

 
		创建型模式：主要用于创建对象
			抽象工厂模式
			建造者模式
			工厂方法模式
			原型模式
			单例模式

		结构型模式：用于处理类和对象的组合
			桥接模式
			组合模式
			装饰模式
			外观模式
			享元模式
			代理模式
			适配器模式

		行为型模式：用于描述 类或对象 怎样交互和怎样分配职责
			职责链模式
			命令模式
			解释器模式
			迭代模式
			中介者模式
			备忘录模式
			观察者模式
			状态模式
			策略模式
			模板方法模式
			访问者模式
			
	4.编程范式
		并发编程：
			异步模式
			并发控制
			并发模型
			并发io
		函数式编程
		面向对象编程
		响应式编程
	
	5.编译原理    轮子哥推荐《Parsing Techniques》   后端用Engineering a Compiler
		编译器 ： 编译器就是将  高级语言  翻译为另一种 低级语言 的程序
				 流程：源代码 (source code) → 预处理器 (preprocessor) → 编译器 (compiler) → 目标代码 (object code) → 链接器 (Linker) → 可执行程序 (executables)
		formalsemantics
		内存管理
		类型理论
	
	
	6.软件开发
	敏捷开发
	技术管理
	重构
	版本控制
	
	7.软件测试

编程语言

	1.网络
		网络管理
		套接字
		http
	websocket
	2.存储
	    缓存
	    数据库:MySQL  redis  mongodb
	    文件系统:文件寻址  文件监控
	3.语法基础
	    /类与对象
	          对象: this   单例模式
	         类继承
	    /流程控制：运算符   分支  循环  迭代器与生成器   异常处理
	    /数据结构
	        变量与作用域:  闭包
	        字符串类型
	        基本类型：数值类型   空   布尔 枚举
	        序列类型：数组  集合
	        索引类型:  映射
	    /函数
	        函数参数
	        函数调用
	        匿名函数
	        装饰器
	    /输入输出
	    /模块
	        
	4.系统进程
		aop切面编程
		并发编程：
			线程：线程池 协程
			并发控制： 锁  锁与事务 数据一致性
			异步模式：callback   
					 promise/future/defer
					 generator
					 async/await
		
		系统调用
		响应式编程：rxextension  actor streaAPI
		本地跨语言调用
	
	5.进阶
		元编程： 反射与代理 数据绑定 代码生成 依赖注入/控制反转
		泛型编程
		内存管理： 内存结构 内存分配 垃圾回收
		工程实践： 代码性能
				  样式指南：
				  		  代码风格约定
				  		  命名约定
				  		  项目/模块架构
				  		  语法检查
				  		  文档注释
		设计模式：SOLID  Observable 	
		算法实现	  
	6.测试与发布
		调试
		测试：断言  mock&stub  testrunner 
		发布
		日志


web前端
	1.前端框架
	mvvm数据双向绑定：angular.js
	工具：jquery lodash
	视图： react vueJS
	状态管理： redux mobx
		
	2.构建辅助
		任务执行：grunt gulp 
		模块打包： JSPM webpack rollup
	
	3.webUI
		动画： css动画  js动画
		数据可视化： 表图  流程图  地图 绘图基础
	4.交互部件
		对话框
		画廊
		指示器
		输入器
		表格网格
	5.浏览器原理
		自动化浏览器：jsdom  selenium phantomjs
		浏览器引擎： electron nw.js
	6.前端进阶
		前端工程化： 组件化 模块化 状态管理 同构
		前端衍化
		测试与发布：调试
				   测试：
				   		测试执行器
				   		端到端测试
				   		断言库
				   发布
		性能优化：资源加载 关键路径 渲染优化
	
	7.语法基础
		dom：事件响应 网络请求 存储 sysproc 
		css：装饰 布局 预处理器 后处理器
		html:语义化布局
		webassembly





	
工作
    前端

    浏览器 js HTML css sublime  标准规范  
    前端框架 oop 数据结构闭包 设计模式 
    编译 协议 操作系统 算法 移动响应式  
    类库 模块 可视化 安全 性能 部署 SEO 代码质量测试
    
    后端  https://github.com/xingshaocheng/architect-awesome
    
    架构师
    
    系统架构能力  工具架构能力 架构性能优化  软技能  总结归纳 项目管理
    
    
    安全元
    基础 
        http 抓包调试 跳转 算法 linux  正则  
    web安全
        web服务组件  安全维度 web安全标准 浏览器 http代理 漏洞扫描 漏洞利用 抓包工具 kali
    移动安全
    系统安全
    嵌入式安全
        路由安全 摄像头安全 工控安全 
        
    DBA  数据库管理员

    数据库技能 网络协议 软技能 
     运维部署监控  linux  硬件  
    
    
    
    运维

    linux基础 运维命令  脚本 kvm虚拟化 
        硬件 安全 网络 底层 
    
    基础服务
        LAMP指的Linux（操作系统）、ApacheHTTP 服务器，MySQL（有时也指MariaDB，数据库软件） 和PHP（有时也是指Perl或Python） 的第一个字母，一般用来建立web应用平台 搭建动态网站
        
        DHCP（ 动态主机配置协议）是一个局域网的网络协议，使用UDP协议工作， 主要作用是集中的管理、分配IP地址
        
    
    运维平台工具
        
        ssh证书 
        SSH 为建立在 --应用层-- 基础上的安全协议。SSH 是目前较可靠，
        专为  远程登录 会话和其他网络服务提供  --安全性的协议--
        
        fabric  
        Fabric 是一个 Python (2.5-2.7) 的库和命令行工具，用来提高基于 SSH 的应用部署和系统管理效率。
        
        ansible
        是新出现的自动化运维工具，基于模块工作 
        实现了批量系统配置、批量程序部署、批量运行命令等功能。
    


互联网已经成为世界上最大的计算机网络

互联网包含因特网，因特网包含万维网

	互联网发展的业务   ISP	ICP
	   ISP（互联网服务提供商） 
		指向广大用户综合提供互联网接入业务、信息业务和增值业务的电信运营商（比如中国电信、移动、联通提供的宽度服务）
	ICP（互联网内容提供商）  
		向广大用户综合提供互联网信息和增值业务的企业（比如：新浪、搜狐、腾讯）。其实，我们说的ICP域名备案就是指互联网内容提供商向工信部提交经营审核



因特网 INTERNET是由于许多小的网络（子网）互联而成的一个逻辑网
	是一组全球信息资源的总汇
	
	所有计算机 +电缆 =因特网

	以相互交流信息资源为目的





常用小框架

    负载均衡    LV5
    缓存        tair
    调度系统    tbschedule
    文件存储    oss tbfs
    服务化      dubbo
    分库分表    tddl  cobar
    容器        潘多拉


数据库

    组成：
        数据库：物理操作文件系统或其他形式文件类型的集合；
        数据库实例：MySQL 数据库由后台线程以及一个共享内存区组成；
    
    关系：
        在 MySQL 中，实例和数据库往往都是一一对应的，
        而我们也无法直接操作数据库，而是要通过数据库实例来操作数据库文件，
        可以理解为数据库实例是数据库为上层提供的一个专门用于操作的接口
    
        在 Unix 上，启动一个 MySQL 实例往往会产生两个进程， mysqld 就是真正的数据库服务守护进程，而 mysqld_safe是一个用于检查和设置 mysqld 启动的控制程序，它负责监控 MySQL 进程的执行，当 mysqld 发生错误时， mysqld_safe 会对其状态进行检查并在合适的条件下重启。
    
    1.连接 /线程
    2.分析 缓存 优化
    3.存储
    
    一层用于连接、线程处理的部分并不是 MySQL 『发明』的，很多服务都有类似的组成部分；
    二层中包含了大多数 MySQL 的核心服务，包括了对 SQL 的解析、分析、优化和缓存等功能，存储过程、触发器和视图都是在这里实现的；
    三层就是 MySQL 中真正负责数据的存储和提取的存储引擎，例如：InnoDB、MyISAM 等，文中对存储引擎的介绍都是对 InnoDB 实现的分析。
    
    InnoDB
        在 InnoDB 存储引擎中，所有的数据都被逻辑地存放在表空间中，表空间（tablespace）是存储引擎中最高的存储逻辑单位，在表空间的下面又包括段（segment）、区（extent）、页（page）







架构

 
表现层、数据持久层、业务逻辑层、领域模型层

M-V-X   是展现层的模式，它们消费了业务层的产品  Model, 并且使用 X 来控制用户输入与业务层产品的数据，做出各种新作为参数提交给业务层 对业务进行影响
X里面的逻辑   都是界面交互逻辑
X里面的的调用 都是用户操作调用
X里面的模型， 都是视图模型（VM）

X可以是  C 可以是VM 可以是P





mvc   https://draveness.me/mvx     主页 https://draveness.me/index
	在 GUI 编程领域   mvc 50岁了
	
	设计 MVC 的重要目的就是在人的心智模型与计算机的模型之间建立一个桥梁，
	而 MVC 能够解决这一问题并为用户提供直接看到信息和操作信息的功能。
	
	
	v  视图：管理作为位图展示到屏幕上的图形和文字输出；
	c  控制器：翻译用户的输入并依照用户的输入操作模型和视图；
	m  模型：管理应用的行为和数据，响应数据请求（经常来自视图）和更新状态的指令（经常来自控制器）；
	
	协作方式
		控制器负责对模型中的数据进行更新，而视图向模型中请求数据；
		当有用户的行为触发操作时，会有控制器更新模型，并通知视图进行更新，在这时视图向模型请求新的数据，
		而这就是作者所理解的标准 MVC 模式下，Model、View 和 Controller 之间的协作方式。
	
	依赖关系
		模型层可以单独工作，而视图层和控制器层都依赖与模型层中的数据
	
	核心问题
		MVC 最重要的概念就是分离展示层 Separated Presentation，如何在领域对象（Domain Object）和我们在屏幕上看到的 GUI 元素进行划分是 MVC 架构模式中最核心的问题。
	
	展示层和领域层关系
		GUI 应用程序由于其需要展示内容的特点，分为两个部分：一部分是用于展示内容的展示层（Presentation Layer），另一部分包含领域和数据逻辑的领域层（Domain Layer）。
		展示层依赖于领域层中存储的数据，而领域层对于展示层一无所知，领域层其实也是 MVC 模式中的模型层，而展示层可以理解为 VC 部分。
		MVC 最重要的目的并不是规定各个模块应该如何交互和联系，而是将原有的混乱的应用程序划分出合理的层级，把一团混乱的代码，按照展示层和领域层分成两个部分；在这时，领域层中的领域对象由于其自身特点不需要对展示层有任何了解，可以同时为不同的展示层工作。
	


cms




项目构建工具
	buildout的是一款自动化构建工具。
	由Zope团队开发维护。包名为zc.buildout。
	
	buildout可以为应用构建独立的依赖环境。类似于virtualenv，但二者还有不同。
	粗略地讲，buildout支持的功能更多更便于自动化而且具体定位有所不同。
	知乎使用了buildout作为python项目的构建工具



	VirtualEnv用于在一台机器上创建多个独立的Python虚拟运行环境，多个Python环境相互独立，互不影响，它能够：1.在没有权限的情况下安装新套件 2.不同应用可以使用不同的套件版本 3.套件升级不影响其他应用
	虚拟环境非常有用，因为它可以防止系统出现包管理混乱和版本冲突的问题  虚拟环境不需要管理员权限。	

项目部署
	
	Fabric 是一个 Python (2.5-2.7) 的库和命令行工具，用来提高基于 SSH 的应用部署和系统管理效率。

	ansible是新出现的自动化运维工具，基于模块工作 
	实现了批量系统配置、批量程序部署、批量运行命令等功能。
网络框架

	Twisted
	是用Python实现的基于事件驱动的网络引擎框架

	过程：一个Twisted程序由reactor发起的主循环和一些回调函数组成。
	当事件发生了，比如一个client连接到了server，
	这时候服务器端的事件会被触发执行

   使用 Twisted 的好处在于，它是以事件驱动编程实现的，
   所以提供了事件注册的回调函数的接口，每次接受到请求，
   获得了事件通知，就调用事件所注册的回调函数
   ( Node.js 程序员可能比较熟悉)。
   这让我不必去操心服务器事件驱动的编写。
	并且，在网络引擎方面，有心跳包和粘包的三方库，非常完善。
	不善于处理异步 用Gevent

	事件驱动编程是一种编程范式，这里程序的执行流由外部事件来决定。它的特点是包含一个事件循环，当外部事件发生时使用回调机制来触发相应的处理
	
	 
	Tornado
	异步非阻塞IO的Python Web框架   用于Web服务器

	完备的Web框架：与Django、Flask等一样，Tornado也提供了URL路由映射、Request上下文、基于模板的页面渲染技术等开发Web应用的必备工具。
	是一个高效的网络库，性能与Twisted、Gevent等底层Python框架相媲美：提供了异步I/O支持、超时事件处理。这使得Tornado除了可以作为Web应用服务器框架，还可以用来做爬虫应用、物联网关、游戏服务器等后台应用。
	提供高效HTTPClient：除了服务器端框架，Tornado还提供了基于异步框架的HTTP客户端。
	提供高效的内部HTTP服务器：虽然其他Python网络框架（Django、Flask）也提供了内部HTTP服务器，但它们的HTTP服务器由于性能原因只能用于测试环境。而Tornado的HTTP服务器与Tornado异步调用紧密结合，可以直接用于生产环境。
	完备的WebSocket支持：WebSocket是HTML5的一种新标准，实现了浏览器与服务器之间的双向实时通信。
		
	Gevent 
	gevent 是一种基于协程的 Python 网络库，它用到 greenlet 提供的，封装了 libevent 事件循环的高层同步API。
	协程就是由程序员自己编码实现调度的多线程。
	
	flask
	
	Django
 
	Django是遵循MVC架构的Web开发框架，其主要由以下几部分组成。
	
	管理工具（Management）：一套内置的创建站点、迁移数据、维护静态文件的命令工具。
	模型（Model）：提供数据访问接口和模块，包括数据字段、元数据、数据关系等的定义及操作。
	视图（View）：Django的视图层封装了HTTP Request和Response的一系列操作和数据流，其主要功能包括URL映射机制、绑定模板等。
	模板（Template）：是一套Django自己的页面渲染模板语言，用若干内置的tags和filters定义页面的生成方式。
	表单（Form）：通过内置的数据类型和控件生成HTML表单。
	管理站（Admin）：通过声明需要管理的Model，快速生成后台数据管理网站。

	爬虫框架

		Scrapy 
		Python开发的一个快速、高层次的屏幕抓取和web抓取框架，用于抓取web站点并从页面中提取结构化的数据。Scrapy用途广泛，可以用于数据挖掘、监测和自动化测试。




ORM框架
 	对象-关系映射  面向对象的 对象模型 和 关系型数据 之间的相互转换。
 	基于关系型数据库的数据存储，实现一个虚拟的面向对象的数据访问接口
	
	对象和关系数据是业务实体的两种表现形式，
	业务实体在内存中表现为对象，在数据库中表现为关系数据。
	内存中的对象之间存在关联和继承关系，
	而在数据库中，关系数据无法直接表达多对多关联和继承关系。
	因此，对象-关系映射(ORM)系统一般以中间件的形式存在，
	主要实现程序对象到关系数据库数据的映射。

	4层
	1.表现层
	2.业务逻辑层
	3.持久化层
	4.数据库层





	Django 的 ORM   相比 SQLAlchemy， Django 的 ORM 更吻合于直接操作SQL对象，操作暴露了简单直接映射数据表和Python类的SQL对象 
	SQLAlchemy是Python编程语言下的一款ORM框架，该框架建立在数据库API之上，使用关系对象映射进行数据库操作
	SQLObject 是一个介于SQL数据库和Python之间映射对象的Python ORM
	Storm 是一个介于 单个或多个数据库与Python之间 映射对象的 Python ORM
	peewee 是一个小的，表达式的 ORM   专注于极简主义

	Python ORM 之间对比
	
	SQLObject
	优点：
	
	采用了易懂的ActiveRecord 模式
	
	一个相对较小的代码库
	
	缺点：
	
	方法和类的命名遵循了Java 的小驼峰风格
	
	不支持数据库session隔离工作单元
	
	
	Storm
	优点：
	
	清爽轻量的API，短学习曲线和长期可维护性
	
	不需要特殊的类构造函数，也没有必要的基类
	
	缺点：
	
	迫使程序员手工写表格创建的DDL语句，而不是从模型类自动派生
	
	Storm的贡献者必须把他们的贡献的版权给Canonical公司
	
	
	Django ORM
	优点：
	
	易用，学习曲线短
	
	和Django紧密集合，用Django时使用约定俗成的方法去操作数据库
	
	缺点：
	
	不好处理复杂的查询，强制开发者回到原生SQL
	
	紧密和Django集成，使得在Django环境外很难使用
	
	
	peewee
	优点：
	
	Django式的API，使其易用
	
	轻量实现，很容易和任意web框架集成
	
	缺点：
	
	不支持自动化 schema 迁移
	
	多对多查询写起来不直观
	
	
	SQLAlchemy
	优点：
	
	企业级 API，使得代码有健壮性和适应性
	
	灵活的设计，使得能轻松写复杂查询
	
	缺点：
	
	工作单元概念不常见
	
	重量级 API，导致长学习曲线

Excel的扩展工具 
	快
	xlwings：简单强大，可替代VBA   
	win32com：不仅仅是excel，可以处理office;
	慢
	openpyxl：简单易用，功能广泛
	pandas：使用需要结合其他库，数据处理是pandas立身之本
	DataNitro：作为插件内嵌到excel中，可替代VBA，在excel中优雅的使用python
	xlutils：结合xlrd/xlwt，老牌python包，需要注意的是你必须同时安装这三个库


数据分析
	NumPy系统是Python的一种开源的数值计算扩展
	内 容 一个强大的N维数组对象Array 
	用 途 存储和处理大型矩阵
	
	
	pandas 是基于NumPy 的一种工具,该工具是为了解决数据分析任务而创建的
	


不要用正则表达式来解析 HTML

KfrEW7jTMOec

打开网页  webbrowser

import webbrowser
for i in range(10):
    webbrowser.open('http://inventwithpython.com/')





requests--------------------------------------------------------------

检查错误

res = requests.get('http://inventwithpython.com/page_that_does_not_exist')
res.raise_for_status()



原始响应内容
Response.iter_content是可迭代对象

>>> r = requests.get('https://github.com/timeline.json', stream=True)
>>> r.raw
<requests.packages.urllib3.response.HTTPResponse object at 0x101194810>
>>> r.raw.read(10)
'\x1f\x8b\x08\x00\x00\x00\x00\x00\x00\x03'

with open(filename, 'wb') as fd:
    for chunk in r.iter_content(chunk_size):
        fd.write(chunk)






selenium--------------------------------------------------------------------------


from selenium import webdriver
browser = webdriver.Firefox()
browser.get('https://www.baidu.com/')

点击事件
browser.get('https://www.hao123.com/')
link = browser.find_element_by_link_text('爱 奇 艺')
link.click()

browser.back()       点击“返回”按钮。 
browser.forward()    点击“前进”按钮。
browser.refresh()    点击“刷新”按钮。
browser.quit()       点击“关闭窗口”按钮








Excel---------------------------------------------------------------
wb.get_sheet_names(列表) 获得表名
sheet = wb.get_sheet_by_name('Sheet3')  获得表3的名字
sheet.title  表格的title

import openpyxl,os
os.chdir("Z:\\e\\")
wb = openpyxl.load_workbook('51.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheet['A1'].value

 sheet.cell(row=1, column=2).value
 'Apples'
 for i in range(1, 8, 2):
 	print(i, sheet.cell(row=i, column=2).value)



创建保存
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws.append([1, 2, 3])
import datetime
ws['A2'] = datetime.datetime.now()
wb.save("sample.xlsx")

循环行和列
import openpyxl,os
os.chdir(".")
wb = openpyxl.load_workbook('51.xlsx')
sheet = wb.get_sheet_by_name('s')
sheet['A1'] = '123456'
for i in range(1, 8):
 	print(i, sheet.cell(row=i, column=2).value)


打印A1到C3
 import openpyxl,os
 os.chdir(".")
 wb = openpyxl.load_workbook('51.xlsx')
 sheet = wb.get_sheet_by_name('Sheet1')
 for rowOfCellObjects in sheet['A1':'C3']:
     for cellObj in rowOfCellObjects:
         print(cellObj.coordinate, cellObj.value)
 print('--- END OF ROW ---')

打印表中数据
 for row in wb['sheet'].rows:
         print( *[ cell.value for cell in row ])

 for row in wb['sheet'].values:
         print(*row)


openpyxl

Workbook是对工作簿的抽象，Worksheet是对表格的抽象，Cell是对单元格的抽象


Workbook属性

	wb.active      ：获取当前活跃的Worksheet
	wb.worksheets  ：以列表的形式返回所有的Worksheet(表格)
	wb.read_only   ：判断是否以read_only模式打开Excel文档
	wb.encoding    ：获取文档的字符集编码
	wb.properties  ：获取文档的元数据，如标题，创建者，创建日期等
	wb.sheetnames  ：获取工作簿中的表（列表）

	 
	方法
    
	wb = openpyxl.load_workbook('51.xlsx')
	ws = wb['sheet']
	print(ws['A1'].value)
	get_sheet_names 废弃  ---->换成  wb.sheetnames
	get_sheet_by_name    ---->       wb['sheet']

	get_active_sheet    ：获取活跃的表格(新版建议通过active属性获取)
	remove_sheet        ：删除一个表格
	create_sheet        ：创建一个空的表格
	copy_worksheet      ：在Workbook内拷贝表格
	


Worksheet对象 属性
	title              ：表格的标题                                                                                                                                                            
	dimensions         ：表格的大小，这里的大小是指含有数据的表格的大小，即 左上角的坐标:右下角的坐标
	max_row            ：表格的最大行                                                                                                                                                
	min_row            ：表格的最小行                                                                                                                                                
	max_column         ：表格的最大列                                                                                                                                    
	min_column         ：表格的最小列                                                                                                                                    
	rows               ：按行获取单元格(Cell对象) - 生成器                                                                                                
	columns            ：按列获取单元格(Cell对象) - 生成器                                                                                    
	freeze_panes       ：冻结窗格                                                                                                                                    
	values             ：按行获取表格的内容(数据)  - 生成器                                                                                            
	
	方法
	iter_rows：按行获取所有单元格，内置属性有(min_row,max_row,min_col,max_col)
	iter_columns：按列获取所有的单元格
	append：在表格末尾添加数据
	merged_cells：合并多个单元格
	unmerged_cells：移除合并的单元格


cell

	row：单元格所在的行
	column：单元格坐在的列
	value：单元格的值
	coordinate：单元格的坐标	
 print(wb['sheet'].cell(row=1,column=2).coordinate)   B1
 print(wb['sheet'].cell(row=1,column=2).value)        ABC 
 print(wb['sheet'].cell(row=1,column=2).row)          B
 print(wb['sheet'].cell(row=1,column=2).column)       1
 